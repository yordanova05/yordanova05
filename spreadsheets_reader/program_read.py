import openpyxl as xl
from openpyxl.chart import BarChart, Reference

def process_workbook(filename_path):
    wb = xl.load_workbook(filename_path)

    sheet = wb["Sheet1"]
    # cell = sheet["a1"]
    # # cell = sheet.cell(1,1)
    # print(cell.value)

    print(sheet.max_row)

    for row in range(2, sheet.max_row + 1):
        cell_s = sheet.cell(row,3)
        new_cell = cell_s.value*0.9
        corrected_price = sheet.cell(row,4)
        corrected_price.value = new_cell

    values = Reference(sheet, 
            min_row=2, 
            max_row=sheet.max_row, 
            min_col=4,
            max_col=4)

    chart = BarChart()
    chart.add_data(values)
    sheet.add_chart(chart,"h3")

    wb.save(filename_path)


name = input("Enter path: ")
process_workbook(name)